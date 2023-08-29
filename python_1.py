from openpyxl import Workbook

input_file = "input.txt"
data ={}

with open(input_file,"r") as file:
        check = False
        for line in file:
                if "Serving Cell Measurements Response" in line:
                    check = True
                elif check:
                    if "Time :" in line:
                        key,value = line.strip().split(" : ")
                        data[key] = value
                    if "PCI" in line:
                        key,value = line.strip().split(" : ")
                        data[key] = value
                    if "RSRP" in line:
                        key,value = line.strip().split(" : ")
                        data[key] = value
                
print(data)

out_file = "out.xlsx"

wb = Workbook()
ws = wb.active
# Write data to the worksheet
ws.append(["Time", "PCI", "RSRP", "SrxLev", "Rank"])
for row, (key,value) in enumerate(data.items(),start=1):
    # ws.cell(row=1, column=row, value=key)
    ws.cell(row=2, column=row, value=value)

    
wb.save(out_file)