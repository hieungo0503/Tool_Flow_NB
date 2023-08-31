from openpyxl import Workbook

input_file = "input.txt"
data_Cell_S = []
data_Cell_N = []
data_cell_Rank =[]
data_Cell_Start = []
data_Cell_Fail = []

with open(input_file, "r") as file:
    current_entry = {}
    check_1=check_2=check_3=check_4=check_5 = False
    for line in file:
        line = line.strip()
        if "***" in line:
            check_1=check_2=check_3=check_4=check_5 = False
        if "Serving Cell Measurements Response" in line:
            check_1 = True
        elif "Neighbor Cell Measurements" in line:  
            check_2 = True
        elif "Reselection Started Event" in line:
            check_4 = True
        elif "Reselection Failure Event" in line:  
            check_5 = True
        elif "Cell Reselection" in line:      #for RANK  
            check_3 = True
        
        
        if check_1:
            if line.startswith("Time :"):
                current_entry["Time"] = line.split(" : ")[1].strip()
            elif "PCI :" in line:
                current_entry["PCI"] = line.split(":")[1].strip()
            elif "RSRP :" in line:
                current_entry["RSRP"] = line.split(":")[1].strip()
            elif "SrxLev :" in line:
                current_entry["SrxLev"] = line.split(":")[1].strip()
            elif "Rank :" in line:
                current_entry["Rank"] = line.split(":")[1].strip()
                data_Cell_S.append(current_entry)
                current_entry = {}
            # elif "***" in line:
            #     if "Neighbor Cell Measurements" in line:  
            #         check_2 = True
            #     elif "Cell Reselection" in line:      #for RANK  
            #         check_3 = True
            #     check_1 = False
        if check_2:
            if "Time :" in line:
                current_entry["Time"] = line.split(" : ")[1].strip()
            elif "Number of Cells" in line:
                current_entry["Number of Cells"] = line.split(":")[1].strip()
            elif "PCI :" in line:
                current_entry["PCI"] = line.split(":")[1].strip()
            elif "RSRP :" in line:
                current_entry["RSRP"] = line.split(":")[1].strip()
            elif "RSRQ :" in line:
                current_entry["RSRQ"] = line.split(":")[1].strip()
                data_Cell_N.append(current_entry)
                current_entry = {}
            # elif "***" in line:
            #     if "Serving Cell Measurements Response" in line:
            #          check_1 = True
            #     elif "Cell Reselection" in line:      #for RANK  
            #          check_3 = True
            #     check_2 = False
        if check_3:
            if "Time :" in line:
                current_entry["Time"] = line.split(" : ")[1].strip()
            elif "Number of Cells" in line:
                current_entry["Number of Cells"] = line.split(":")[1].strip()
            elif "PCI :" in line:
                current_entry["PCI"] = line.split(":")[1].strip()
            elif "Rank :" in line:
                current_entry["Rank"] = line.split(":")[1].strip()
            elif "TReselection Value :" in line:
                current_entry["TReselection Value"] = line.split(":")[1].strip()
                data_cell_Rank.append(current_entry)
                current_entry = {}
            # elif "***" in line:
            #     if "Serving Cell Measurements Response" in line:
            #         check_1 = True
            #     elif "Neighbor Cell Measurements" in line:  
            #         check_2 = True
            #     check_3 = False
        if check_4:
            if "Time :" in line:
                current_entry["Time"] = line.split(" : ")[1].strip()
            elif "PCI :" in line:
                current_entry["PCI"] = line.split(":")[1].strip()
                data_Cell_Start.append(current_entry)
                current_entry = {}
        if check_5:
            if "Time :" in line:
                current_entry["Time"] = line.split(" : ")[1].strip()
            elif "PCI :" in line:
                current_entry["PCI"] = line.split(":")[1].strip()
            elif "CellReselFailureCause :" in line:
                current_entry["CellReselFailureCause"] = line.split(":")[1].strip()
                data_Cell_Fail.append(current_entry)
                current_entry = {}
            

# Reselection Started Event
# Reselection Failure Event CellReselFailureCause
            
        
# Print collected data_Cell_S (for verification)
# for entry in data_Cell_S:
#     print(entry)

out_file = "out_S_cell.xlsx"

wb = Workbook()
ws = wb.active

# Write headers
ws.append(["Time", "PCI", "RSRP", "SrxLev", "Rank"])

# Write data_Cell_S to the worksheet
for entry in data_Cell_S:
    ws.append([
        entry.get("Time", ""),
        entry.get("PCI", ""),
        entry.get("RSRP", ""),
        entry.get("SrxLev", ""),
        entry.get("Rank", "")
    ])

wb.save(out_file)
print("data_Cell_S has been written to", out_file)

# for entry in data_Cell_N:
#     print(entry)

out_file = "out_N_cell.xlsx"
wb = Workbook()
ws = wb.active
ws.append(["Time","Number of Cells","PCI","RSRP","RSRQ"])
for entry in data_Cell_N:
    ws.append([
        entry.get("Time",""),
        entry.get("Number of Cells",""),
        entry.get("PCI",""),
        entry.get("RSRP",""),
        entry.get("RSRQ","")
    ])
wb.save(out_file)
print("Write data to N_cell OK")

# for entry in data_cell_Rank:
#     print(entry)

out_file = "out_Rank.xlsx"
wb = Workbook()
ws = wb.active
ws.append(["Time","Number of Cells","PCI","Rank","TReselection Value"])
for entry in data_cell_Rank:
    ws.append([
        entry.get("Time",""),
        entry.get("Number of Cells",""),
        entry.get("PCI",""),
        entry.get("Rank",""),
        entry.get("TReselection Value","")
    ])
wb.save(out_file)
print("Write data to Rank OK")


for entry in data_Cell_Start:
    print(entry)

out_file = "out_Cell_Start.xlsx"
wb = Workbook()
ws = wb.active
ws.append(["Time","PCI"])
for entry in data_Cell_Start:
    ws.append([
        entry.get("Time",""),
        entry.get("PCI","")
        
    ])
wb.save(out_file)
print("Write data to Start OK")


for entry in data_Cell_Fail:
    print(entry)

out_file = "out_Fail.xlsx"
wb = Workbook()
ws = wb.active
ws.append(["Time","PCI","CellReselFailureCause"])
for entry in data_Cell_Fail:
    ws.append([
        entry.get("Time",""),
        entry.get("PCI",""),
        entry.get("CellReselFailureCause","")
    ])
wb.save(out_file)
print("Write data to Fail OK")